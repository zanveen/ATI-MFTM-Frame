import streamlit as st
import json
import copy
import os
import base64
import pandas as pd
import gspread
import zipfile
import calendar
from google.oauth2.service_account import Credentials
from datetime import datetime, date, timedelta
from pathlib import Path
from openpyxl import load_workbook
import io

from PIL import Image as PILImage

# ─── 설정 ───
_logo_icon = PILImage.open("ati_logo.png") if os.path.exists("ati_logo.png") else "🏭"
st.set_page_config(
    page_title="Frame 제작 현황 관리",
    page_icon=_logo_icon,
    layout="wide",
    initial_sidebar_state="expanded"
)

DATA_DIR = Path("data")
DATA_DIR.mkdir(exist_ok=True)
PROJECTS_FILE = DATA_DIR / "projects.json"
TEMPLATE_FILE = Path("template.xlsx")
LOGO_FILE = "ati_logo.png"

# ─── 사용자 권한 정보 ───
USER_CREDENTIALS = {
    "rladbstn5344": {"role": "admin", "name": "관리자", "company": "ALL"},
    "11053": {"role": "vendor", "name": "한울산업", "company": "한울산업"},
    "10738": {"role": "vendor", "name": "정한테크", "company": "정한테크"}
}

# ─── 체크리스트 항목 정의 ───
CHECKLIST_ITEMS = [
    {"no": 1, "category": "판금설계", "item": "수정 도면 반영 확인"},
    {"no": 2, "category": "제작", "item": "용접 방향 확인"},
    {"no": 3, "category": "제작", "item": "비드 상태 확인"},
    {"no": 4, "category": "제작", "item": "주요 치수 정밀 검사 (대각 기장, 폭, 높이 등)"},
    {"no": 5, "category": "제작", "item": "절곡부 마감 상태"},
    {"no": 6, "category": "제작", "item": "SUS FRAME 얼룩 확인"},
    {"no": 7, "category": "제작", "item": "용접 후 뒤틀림 및 열변형 교정 상태 확인"},
    {"no": 8, "category": "도장", "item": "도장일 확인 및 건조 상태"},
    {"no": 9, "category": "도장", "item": "도장면 외관 검사 (오염, 흐름, 뭉침, 찍힘 등)"},
    {"no": 10, "category": "도장", "item": "마스킹 부위 및 탭(Tap) 구멍 도료 유입 여부"},
    {"no": 11, "category": "조립", "item": "도어 미부착 개소 있는지 확인"},
    {"no": 12, "category": "조립", "item": "도어 개폐시 간섭 혹은 소음 확인"},
    {"no": 13, "category": "조립", "item": "최종 조립 수평 상태 및 프레임 단차 확인"},
    {"no": 14, "category": "조립", "item": "명판, 경고 라벨 등 부착물 상태 확인"},
    {"no": 15, "category": "납품", "item": "납품일까지 미입고품 확인 및 전달"},
    {"no": 16, "category": "납품", "item": "레벨풋 장착 확인"},
    {"no": 17, "category": "납품", "item": "프레임 리스트 동봉 확인"},
    {"no": 18, "category": "절곡 단품", "item": "납품일까지 미입고품 확인 및 전달"},
    {"no": 19, "category": "절곡 단품", "item": "표면 스크래치, 찍힘 및 모서리 버(Burr) 제거 상태"},
    {"no": 20, "category": "절곡 단품", "item": "파트별 수량 확인 및 라벨링(식별표) 부착 상태"},
]

# ─── 로고 이미지 함수 ───
def get_logo_html(height="34px"):
    if os.path.exists(LOGO_FILE):
        try:
            with open(LOGO_FILE, "rb") as f:
                encoded = base64.b64encode(f.read()).decode()
            return f'<img src="data:image/png;base64,{encoded}" style="height:{height}; vertical-align:middle; margin-right:8px; margin-bottom:4px;">'
        except: pass
    return "<span style='color:#4A90D9; font-weight:900;'>ATI</span> "

# ─── 구글 시트 연동 및 복구 로직 ───
def get_gspread_client():
    scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
    credentials = Credentials.from_service_account_info(st.secrets["gcp_service_account"], scopes=scope)
    return gspread.authorize(credentials)

def load_from_sheets():
    try:
        client = get_gspread_client()
        sh = client.open("Frame_Data")
        worksheet = sh.sheet1 
        all_records = worksheet.get_all_records()
        
        projects = {}
        for row in all_records:
            pid = str(row.get('pid', '')).strip()
            if not pid: continue 
            
            projects[pid] = {
                "info": {
                    "company": str(row.get('company', '미상')),
                    "equipment": str(row.get('equipment', '미상')),
                    "order_date": str(row.get('order_date', '')),
                    "delivery_date": str(row.get('delivery_date', '')),
                    "frame_parts": int(row.get('frame_parts', 1)) if row.get('frame_parts') else 1,
                    "frame_options": json.loads(row.get('frame_options', '[]')) if row.get('frame_options') else [],
                    "exterior_spec": str(row.get('exterior_spec', '')),
                    "interior_spec": str(row.get('interior_spec', '')),
                    "notes_top": str(row.get('notes_top', '')),
                    "delivery_delay_count": int(row.get('delivery_delay_count', 0)) if row.get('delivery_delay_count') else 0,
                    "delay_total_biz_days": int(row.get('delay_total_biz_days', 0)) if row.get('delay_total_biz_days') else 0,
                    "delay_request": json.loads(row.get('delay_request', '{}')) if row.get('delay_request') else {},
                    "is_delivered": str(row.get('is_delivered', '')).upper() in ('TRUE', '1', 'YES')
                },
                "checks": json.loads(row.get('checks', '{}')) if row.get('checks') else {},
                "special_notes": str(row.get('special_notes', '')),
                "history": json.loads(row.get('history', '[]')) if row.get('history') else []
            }
        return projects
    except Exception as e:
        return {}

def save_to_sheets(projects):
    try:
        client = get_gspread_client()
        sh = client.open("Frame_Data")
        worksheet = sh.sheet1
        
        header = ["pid", "company", "equipment", "order_date", "delivery_date", "frame_parts", 
                  "frame_options", "exterior_spec", "interior_spec", "notes_top", 
                  "delivery_delay_count", "delay_total_biz_days", "delay_request", "checks", "special_notes", "history", "is_delivered"]
        
        data_to_save = [header]
        for pid, p in projects.items():
            if not p.get('info'): continue 
            info = p['info']
            data_to_save.append([
                pid, info.get('company',''), info.get('equipment',''), info.get('order_date',''), info.get('delivery_date',''),
                info.get('frame_parts', 1), json.dumps(info.get('frame_options', []), ensure_ascii=False),
                info.get('exterior_spec',''), info.get('interior_spec',''), info.get('notes_top',''),
                info.get('delivery_delay_count', 0), info.get('delay_total_biz_days', 0),
                json.dumps(info.get('delay_request', {}), ensure_ascii=False),
                json.dumps(p.get('checks', {}), ensure_ascii=False), p.get('special_notes',''), 
                json.dumps(p.get('history', []), ensure_ascii=False), info.get('is_delivered', False)
            ])
        
        worksheet.clear()
        try:
            worksheet.update(values=data_to_save, range_name='A1')
        except TypeError: 
            worksheet.update('A1', data_to_save)
            
    except Exception as e:
        st.error(f"저장 실패: {e}")

# ─── 계산 및 유틸 함수 ───
def repair_project(p):
    if not isinstance(p, dict): p = {}
    if "info" not in p: p["info"] = {}
    info = p["info"]
    info.setdefault("company", "미상")
    info.setdefault("equipment", "알수없음")
    info.setdefault("delivery_date", "")
    info.setdefault("delivery_delay_count", 0)
    info.setdefault("delay_total_biz_days", 0)
    info.setdefault("frame_parts", 1)
    info.setdefault("frame_options", [])
    info.setdefault("is_delivered", False)
    if "delay_request" not in info or not isinstance(info["delay_request"], dict):
        info["delay_request"] = {}
    if "checks" not in p: p["checks"] = {}
    if "history" not in p: p["history"] = []
    if "special_notes" not in p: p["special_notes"] = ""
    return p

def calc_progress(checks):
    total_score = sum(5 if c.get("status") == "확인" else 2 if c.get("status") == "미비" else 0 for c in checks.values())
    return total_score / 100

def calc_score(checks):
    return sum(5 if c.get("status") == "확인" else 2 if c.get("status") == "미비" else 0 for c in checks.values())

def get_project_status(proj):
    if proj.get("info", {}).get("is_delivered", False):
        return "완료"
    
    dd = proj.get("info", {}).get("delivery_date", "")
    if dd:
        try:
            d = datetime.strptime(dd, "%Y-%m-%d").date()
            if (d - date.today()).days <= 7: return "납기임박"
        except: pass
    return "진행중"

def count_business_days(start_date, end_date):
    if not start_date or not end_date: return 0
    try:
        if isinstance(start_date, str): start_date = datetime.strptime(start_date, "%Y-%m-%d").date()
        if isinstance(end_date, str): end_date = datetime.strptime(end_date, "%Y-%m-%d").date()
        days = 0
        current = start_date
        step = timedelta(days=1)
        while current < end_date:
            if current.weekday() < 5:  
                days += 1
            current += step
        return days
    except:
        return 0

def calc_company_delay_stats(all_projects, company, period="전체 누적"):
    total_delays = 0
    total_delay_biz_days = 0
    projs = 0
    for p in all_projects.values():
        info = p.get("info", {})
        if info.get("company") != company: continue
        dd = info.get("delivery_date", "")
        if period != "전체 누적" and not dd.startswith(period): continue
        projs += 1
        total_delays += info.get("delivery_delay_count", 0)
        total_delay_biz_days += info.get("delay_total_biz_days", 0)
    return projs, total_delays, total_delay_biz_days

def filter_projects_by_role(all_projects, user_info):
    if user_info["role"] == "admin": return all_projects
    return {pid: p for pid, p in all_projects.items() if p.get("info", {}).get("company") == user_info["company"]}

def generate_checklist_excel(project):
    wb = load_workbook(TEMPLATE_FILE)
    ws = wb["점검표"]
    info = project.get("info", {})
    checks = project.get("checks", {})

    ws["C2"], ws["C3"], ws["C4"], ws["C5"] = info.get("company", ""), info.get("equipment", ""), info.get("order_date", ""), info.get("delivery_date", "")
    ws["C6"], ws["C7"], ws["C8"] = info.get("exterior_spec", ""), info.get("interior_spec", ""), info.get("frame_parts", "")
    
    frame_options = info.get("frame_options", [])
    notes = info.get("notes_top", "")
    combined_d3 = ""
    if frame_options: combined_d3 += f"[옵션: {', '.join(frame_options)}]\n"
    if notes: combined_d3 += notes
    if combined_d3: ws["D3"] = combined_d3.strip()

    for i, item in enumerate(CHECKLIST_ITEMS):
        row = 13 + i
        key = str(item["no"])
        status = checks.get(key, {}).get("status", "")
        if status == "확인": ws[f"D{row}"] = "○"
        elif status == "미비": ws[f"E{row}"] = "○"
        c_date = checks.get(key, {}).get("date", "")
        if c_date: ws[f"F{row}"] = c_date
        c_memo = checks.get(key, {}).get("memo", "")
        if c_memo: ws[f"H{row}"] = c_memo

    ws["F34"] = f"{calc_score(checks)} / 100" 
    ws["A37"] = project.get("special_notes", "")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# ─── CSS 스타일 ───
st.markdown("""
<style>
    .login-container { max-width: 420px; margin: 80px auto; padding: 40px; background: white; border-radius: 16px; box-shadow: 0 10px 25px rgba(0,0,0,0.1); text-align: center; border-top: 6px solid #4A90D9; }
    
    /* 사이드바 크기 및 폰트 */
    [data-testid="stSidebar"] { min-width: 20vw !important; max-width: 20vw !important; background-color: #f0f4f8; }
    [data-testid="stSidebar"] p, [data-testid="stSidebar"] span { font-size: 18px !important; line-height: 1.8 !important; }
    
    /* 🔥 사이드바 메뉴 탭(박스) 폭을 로그아웃 박스(100%)와 1px 오차 없이 완벽하게 맞춤 🔥 */
    [data-testid="stSidebar"] .stRadio { width: 100% !important; }
    [data-testid="stSidebar"] div[data-testid="stRadio"], 
    [data-testid="stSidebar"] div[role="radiogroup"] { 
        width: 100% !important; 
        display: flex !important;
        flex-direction: column !important;
        align-items: stretch !important; /* 자식 요소 100% 꽉 채우기 */
    }
    [data-testid="stSidebar"] div[role="radiogroup"] { gap: 8px; }
    [data-testid="stSidebar"] div[role="radiogroup"] > label {
        width: 100% !important;
        display: flex !important;
        box-sizing: border-box !important;
        background-color: white;
        border: 1px solid #cbd5e1;
        border-radius: 10px;
        padding: 12px 15px !important;
        margin-bottom: 8px;
        cursor: pointer;
        transition: all 0.2s ease;
        box-shadow: 0 1px 3px rgba(0,0,0,0.05);
    }
    [data-testid="stSidebar"] div[role="radiogroup"] > label > div:first-child { display: none; }
    [data-testid="stSidebar"] div[role="radiogroup"] > label:hover {
        border-color: #4A90D9; transform: translateY(-1px); box-shadow: 0 4px 6px rgba(0,0,0,0.1);
    }
    [data-testid="stSidebar"] div[role="radiogroup"] > label[data-checked="true"] {
        background-color: #4A90D9 !important; border-color: #4A90D9 !important;
    }
    [data-testid="stSidebar"] div[role="radiogroup"] > label[data-checked="true"] p {
        color: white !important; font-weight: 800 !important;
    }
    [data-testid="stSidebar"] div[role="radiogroup"] > label p {
        font-size: 18px !important; font-weight: 600 !important; color: #475569; margin:0; width: 100%; text-align: left;
    }
    
    /* 메인 본문 */
    .main .block-container p, .main .block-container span, .main .block-container label, .main .block-container li, .main .block-container td, .main .block-container th { font-size: 18px !important; line-height: 1.9 !important; }
    .main .block-container h1 { font-size: 34px !important; font-weight: 800; color: #1e293b; }
    
    /* 프로젝트 카드 및 프로그래스바 */
    .project-card { background: white; border-radius: 12px; padding: 22px; margin-bottom: 14px; border-left: 5px solid #4A90D9; box-shadow: 0 2px 8px rgba(0,0,0,0.08); }
    .progress-bar-bg { background: #e8ecf1; border-radius: 10px; height: 26px; overflow: hidden; margin: 12px 0 8px 0; }
    .progress-bar-fill { height: 100%; border-radius: 10px; display: flex; align-items: center; justify-content: center; color: white; font-weight: bold; font-size: 13px; transition: width 0.4s ease; }
    .status-badge { display: inline-block; padding: 4px 12px; border-radius: 20px; font-size: 14px; font-weight: 700; box-shadow: 0 1px 3px rgba(0,0,0,0.1); }
    .badge-green { background: #d4edda; color: #155724; border: 1px solid #c3e6cb; }
    .badge-yellow { background: #fff3cd; color: #856404; border: 1px solid #ffeeba; }
    .badge-red { background: #f8d7da; color: #721c24; border: 1px solid #f5c6cb; }
    .badge-blue { background: #d1ecf1; color: #0c5460; border: 1px solid #bee5eb; }

    /* 대시보드 통짜 버튼 글로벌 CSS 부분은 충돌방지를 위해 삭제하고 아래 반복문 안에 마커로 정밀 주입합니다 */
</style>
""", unsafe_allow_html=True)

# ─── 세션 초기화 및 시트 로드 ───
if "logged_in" not in st.session_state: st.session_state.logged_in = False
if "user_info" not in st.session_state: st.session_state.user_info = None
if "dashboard_filter" not in st.session_state: st.session_state.dashboard_filter = "전체"
if "inspection_project" not in st.session_state: st.session_state.inspection_project = None
if "flash_msg" not in st.session_state: st.session_state.flash_msg = None

if "projects" not in st.session_state:
    st.session_state.projects = load_from_sheets()

# ═══════════════════════════════════════════════════
# 1. 로그인
# ═══════════════════════════════════════════════════
if not st.session_state.logged_in:
    c1, c2, c3 = st.columns([1, 2, 1])
    with c2:
        st.markdown(f"""
        <div class="login-container">
            <div style="font-size: 28px; font-weight: 800; color: #1a1a2e; margin-bottom: 20px;">
                {get_logo_html('34px')} FRAME 제작 통합 관리
            </div>
            <p style="color:#666; margin-bottom:30px;">비밀번호를 입력해주세요.</p>
        """, unsafe_allow_html=True)
        _, pwd_col, _ = st.columns([1, 2, 1])
        with pwd_col:
            pwd = st.text_input("비밀번호", type="password", label_visibility="collapsed")
        st.markdown("<br>", unsafe_allow_html=True)
        login_col1, login_col2, login_col3 = st.columns([1, 2, 1])
        with login_col2:
            if st.button("로그인", use_container_width=True, type="primary"):
                if pwd in USER_CREDENTIALS:
                    st.session_state.logged_in = True
                    st.session_state.user_info = USER_CREDENTIALS[pwd]
                    st.session_state.projects = load_from_sheets() 
                    st.rerun()
                else:
                    st.error("비밀번호가 일치하지 않습니다.")
        st.markdown('</div>', unsafe_allow_html=True)

# ═══════════════════════════════════════════════════
# 2. 메인 애플리케이션
# ═══════════════════════════════════════════════════
else:
    user = st.session_state.user_info

    # ─── 사이드바 ───
    st.sidebar.markdown(f"""
        <div style="font-size:30px; font-weight:bold; margin-bottom:20px; line-height:1.2;">
            {get_logo_html('32px')} FRAME 진척률 관리
        </div>
    """, unsafe_allow_html=True)
        
    st.sidebar.markdown(f"**환영합니다. {user['name']}님**")
    st.sidebar.markdown("---")

    menu_options = ["🏠 HOME", "📅 캘린더", "📋 점검", "📥 양식 추출"]
    if user["role"] == "admin":
        menu_options.insert(1, "➕ 신규 등록")

    menu = st.sidebar.radio("메뉴", menu_options, label_visibility="collapsed")
    if menu != "📋 점검":
        st.session_state.inspection_project = None

    st.sidebar.markdown("---")
    if st.sidebar.button("로그아웃", use_container_width=True):
        st.session_state.logged_in = False
        st.session_state.user_info = None
        st.rerun()
        
    projects = filter_projects_by_role(st.session_state.projects, user)

    if st.session_state.flash_msg:
        st.toast(st.session_state.flash_msg, icon="✅")
        st.success(st.session_state.flash_msg)
        if "등록" in st.session_state.flash_msg:
            st.balloons()
        st.session_state.flash_msg = None 

    # ═══════════════════════════════════════════════════
    # HOME 메뉴
    # ═══════════════════════════════════════════════════
    if menu == "🏠 HOME":
        col_title, col_stats = st.columns([5, 5])
        with col_title:
            st.markdown(f"""
                <div style="font-size:34px; font-weight:800; color:#1e293b; margin-top:10px;">
                    {get_logo_html('34px')} {'전체 진행 현황' if user['role']=='admin' else '우리 회사 배정 프로젝트'}
                </div>
            """, unsafe_allow_html=True)
            
        with col_stats:
            if user["role"] == "admin":
                months = set(p.get("info", {}).get("delivery_date", "")[:7] for p in st.session_state.projects.values() if p.get("info", {}).get("delivery_date"))
                month_list = sorted(list(months), reverse=True)
                selected_period = st.selectbox("통계 조회 기간", ["전체 누적"] + month_list)

        if user["role"] == "admin":
            pending_reqs = [(pid, p) for pid, p in projects.items() if p.get("info", {}).get("delay_request", {}).get("status") == "pending"]
            if pending_reqs:
                req_details = " / ".join([f"<b>{p.get('info',{}).get('equipment','')}</b>({p.get('info',{}).get('company','')})" for _, p in pending_reqs])
                st.markdown(f"""
                <div style="background:#fff3cd; border-left:5px solid #ffc107; padding:10px 16px; border-radius:6px; margin:8px 0;">
                    <b style="color:#856404;">⚠️ 납기 변경 요청 {len(pending_reqs)}건:</b>
                    <span style="font-size:14px; color:#856404;"> {req_details} → '점검' 메뉴에서 승인/반려</span>
                </div>
                """, unsafe_allow_html=True)

        if user["role"] == "admin":
            h_projs, h_delays, h_biz = calc_company_delay_stats(st.session_state.projects, "한울산업", selected_period)
            j_projs, j_delays, j_biz = calc_company_delay_stats(st.session_state.projects, "정한테크", selected_period)

            st.markdown(f"""
            <div style="background:#f8f9fa; border:1px solid #dee2e6; padding:12px 16px; border-radius:8px; margin-bottom:12px;">
                <div style="display:flex; gap:12px;">
                    <div style="flex:1; background:white; padding:10px; border-radius:6px; border:1px solid #e0e0e0; text-align:center;">
                        <b style="color:#4A90D9; font-size:15px;">한울산업</b><br>
                        <span style="font-size:13px;">프로젝트: <b>{h_projs}</b>건 │ 
                            <span style="color:#e74c3c;">지연: <b>{h_delays}</b>회</span> │ 
                            <span style="color:#e74c3c;">지연일: <b>{h_biz}</b>영업일</span>
                        </span>
                    </div>
                    <div style="flex:1; background:white; padding:10px; border-radius:6px; border:1px solid #e0e0e0; text-align:center;">
                        <b style="color:#4A90D9; font-size:15px;">정한테크</b><br>
                        <span style="font-size:13px;">프로젝트: <b>{j_projs}</b>건 │ 
                            <span style="color:#e74c3c;">지연: <b>{j_delays}</b>회</span> │ 
                            <span style="color:#e74c3c;">지연일: <b>{j_biz}</b>영업일</span>
                        </span>
                    </div>
                </div>
            </div>
            """, unsafe_allow_html=True)

        if not projects:
            st.info("표시할 프로젝트가 없습니다.")
        else:
            completed_pids = [pid for pid, p in projects.items() if get_project_status(p) == "완료"]
            urgent_pids = [pid for pid, p in projects.items() if get_project_status(p) == "납기임박"]
            in_progress_pids = [pid for pid, p in projects.items() if get_project_status(p) == "진행중"]

            current_filter = st.session_state.dashboard_filter
            filters_data = [
                ("전체", len(projects), "전체 프로젝트"),
                ("진행중", len(in_progress_pids), "진행중"),
                ("완료", len(completed_pids), "완료"),
                ("납기임박", len(urgent_pids), "납기임박"),
            ]
            
            # 🔥 대시보드 4개 버튼: 마커를 사용한 정밀 타겟팅 (다른 버튼 오작동 원천 차단, 볼드 가운데 정렬 100% 보장) 🔥
            metric_cols = st.columns(4)
            for i, (key, num, label) in enumerate(filters_data):
                with metric_cols[i]:
                    marker_class = f"dash_btn_mk_{i}"
                    st.markdown(f"<div class='{marker_class}' style='display:none;'></div>", unsafe_allow_html=True)
                    
                    is_active = (current_filter == key)
                    bg_col = "#fef2f2" if (is_active and key == "납기임박") else "#eff6ff" if is_active else "white"
                    border_col = "#e74c3c" if (is_active and key == "납기임박") else "#4A90D9" if is_active else "#e2e8f0"
                    num_col = "#e74c3c" if key == "납기임박" else "#000000"
                    txt_col = "#e74c3c" if key == "납기임박" else "#475569"
                    
                    st.markdown(f"""
                    <style>
                    div.element-container:has(.{marker_class}) + div.element-container button,
                    div.stElementContainer:has(.{marker_class}) + div.stElementContainer button {{
                        height: 120px !important;
                        width: 100% !important;
                        border-radius: 12px !important;
                        background-color: {bg_col} !important;
                        border: 3px solid {border_col} !important;
                        box-shadow: {'0 4px 10px rgba(0,0,0,0.1)' if is_active else '0 2px 4px rgba(0,0,0,0.05)'} !important;
                        padding: 0 !important;
                        transition: transform 0.2s, box-shadow 0.2s !important;
                    }}
                    div.element-container:has(.{marker_class}) + div.element-container button:hover,
                    div.stElementContainer:has(.{marker_class}) + div.stElementContainer button:hover {{
                        transform: translateY(-3px) !important;
                        box-shadow: 0 8px 16px rgba(0,0,0,0.1) !important;
                    }}
                    div.element-container:has(.{marker_class}) + div.element-container button p::first-line,
                    div.stElementContainer:has(.{marker_class}) + div.stElementContainer button p::first-line {{
                        color: {num_col} !important;
                        font-size: 42px !important;
                        font-weight: 900 !important;
                        line-height: 1.3 !important;
                    }}
                    div.element-container:has(.{marker_class}) + div.element-container button p,
                    div.stElementContainer:has(.{marker_class}) + div.stElementContainer button p {{
                        color: {txt_col} !important;
                        font-size: 17px !important;
                        font-weight: 700 !important;
                        text-align: center !important;
                        width: 100% !important;
                        margin: 0 !important;
                        white-space: pre-wrap !important;
                    }}
                    </style>
                    """, unsafe_allow_html=True)
                    
                    # 여기서 type="primary"를 줌으로써 디자인 우선순위 확보
                    if st.button(f"{num}\n{label}", key=f"filter_{key}", use_container_width=True, type="primary"):
                        st.session_state.dashboard_filter = key
                        st.rerun()

            st.markdown("---")

            show_pids = list(projects.keys()) if current_filter == "전체" else in_progress_pids if current_filter == "진행중" else completed_pids if current_filter == "완료" else urgent_pids

            if not show_pids:
                st.info("해당하는 프로젝트가 없습니다.")
            else:
                sort_option = st.selectbox("정렬", ["납품 예정일순", "진척률 낮은순", "진척률 높은순"], label_visibility="collapsed")
                sorted_pids = show_pids.copy()
                if sort_option == "납품 예정일순": sorted_pids.sort(key=lambda x: projects[x].get("info", {}).get("delivery_date", "9999"))
                elif sort_option == "진척률 낮은순": sorted_pids.sort(key=lambda x: calc_progress(projects[x].get("checks", {})))
                else: sorted_pids.sort(key=lambda x: calc_progress(projects[x].get("checks", {})), reverse=True)

                for pid in sorted_pids:
                    proj = projects.get(pid)
                    if not proj: continue
                    
                    info = proj.get("info", {})
                    checks = proj.get("checks", {})
                    pct = int(calc_progress(checks) * 100)
                    score = calc_score(checks)
                    delay_cnt = info.get("delivery_delay_count", 0)
                    is_deliv = info.get("is_delivered", False)

                    if is_deliv: badge, bar_color = '<span class="status-badge" style="background:#e2e8f0; color:#000000; border:1px solid #cbd5e1;">✅ 납품완료</span>', "#95a5a6"
                    elif pct >= 50: badge, bar_color = '<span class="status-badge badge-blue">[진행]</span>', "#4A90D9"
                    elif pct > 0: badge, bar_color = '<span class="status-badge badge-yellow">[제작]</span>', "#f39c12"
                    else: badge, bar_color = '<span class="status-badge badge-red">[대기]</span>', "#e74c3c"

                    dd = info.get("delivery_date", "")
                    days_text = ""
                    if dd and not is_deliv:
                        try:
                            diff = (datetime.strptime(dd, "%Y-%m-%d").date() - date.today()).days
                            if diff < 0: days_text = f'<span style="color:#e74c3c;font-weight:bold;">+{abs(diff)}일</span>'
                            elif diff == 0: days_text = '<span style="color:#e74c3c;font-weight:bold;">오늘</span>'
                            elif diff <= 7: days_text = f'<span style="color:#e74c3c;font-weight:bold;">D-{diff}</span>'
                            else: days_text = f'<span style="color:#888;">D-{diff}</span>'
                        except: pass

                    opts_html = f'<span style="background:#eff6ff;color:#3b82f6;padding:2px 8px;border-radius:12px;font-size:12px;font-weight:700;border:1px solid #bfdbfe;margin-left:8px;">{", ".join(info.get("frame_options", []))}</span>' if info.get("frame_options") else ''
                    delay_text = f" │ <span style='color:#e74c3c;'>지연{delay_cnt}회</span>" if delay_cnt > 0 else ""

                    card_col, btn_col = st.columns([9, 1]) if (not is_deliv and user["role"] == "admin") else (st.columns([10])[0], None)
                    
                    with card_col:
                        st.markdown(f"""
                        <div style="background:white; border-radius:8px; padding:10px 16px; margin-bottom:4px; border-left:4px solid {bar_color}; box-shadow:0 1px 3px rgba(0,0,0,0.05);">
                            <div style="display:flex; justify-content:space-between; align-items:center;">
                                <div style="display:flex; align-items:center; gap:6px;">
                                    <b style="font-size:16px;">{info.get("equipment", pid)}</b>{opts_html}
                                </div>
                                <div style="display:flex; align-items:center; gap:6px; font-size:13px;">{badge} {days_text}</div>
                            </div>
                            <div style="display:flex; justify-content:space-between; align-items:center; margin-top:3px;">
                                <span style="font-size:12px; color:#888;">{info.get("company","-")} │ {dd} │ {score}/100점{delay_text}</span>
                                <b style="font-size:13px; color:{bar_color};">{pct}%</b>
                            </div>
                            <div style="background:#eee; border-radius:3px; height:5px; margin-top:4px; overflow:hidden;">
                                <div style="height:100%; width:{max(pct,2)}%; background:{bar_color}; border-radius:3px;"></div>
                            </div>
                        </div>
                        """, unsafe_allow_html=True)
                    
                    if btn_col and not is_deliv and user["role"] == "admin":
                        with btn_col:
                            if st.button("✅ 납품완료", key=f"btn_done_{pid}", use_container_width=True, help="납품 완료 처리"):
                                st.session_state.projects[pid]["info"]["is_delivered"] = True
                                save_to_sheets(st.session_state.projects)
                                st.session_state.flash_msg = f"✅ [{info.get('equipment')}] 납품 처리 완료!"
                                st.rerun()

    # ═══════════════════════════════════════════════════
    # 📅 캘린더
    # ═══════════════════════════════════════════════════
    elif menu == "📅 캘린더":
        
        @st.dialog("프로젝트 상세 현황", width="large")
        def show_project_details_dialog(date_str, day_projs):
            st.markdown(f"#### 📅 {date_str} 프로젝트 현황 ({len(day_projs)}건)")
            day_projs.sort(key=lambda x: x[1].get('info', {}).get('company', ''))
            
            for pid_item, p in day_projs:
                sp_info = p.get("info", {})
                sp_checks = p.get("checks", {})
                sp_pct = int(calc_progress(sp_checks) * 100)
                sp_score = calc_score(sp_checks)
                sp_deliv = sp_info.get("is_delivered", False)
                sp_dd = sp_info.get("delivery_date", "")
                
                try: sp_diff = (datetime.strptime(sp_dd, "%Y-%m-%d").date() - date.today()).days
                except: sp_diff = 99
                
                if sp_deliv: sp_status_txt = "납품완료"
                elif sp_diff < 0: sp_status_txt = f"납기 {abs(sp_diff)}일 초과"
                elif sp_diff <= 7: sp_status_txt = f"D-{sp_diff}"
                else: sp_status_txt = f"D-{sp_diff}"
                
                opts_str = f"  [옵션: {', '.join(sp_info.get('frame_options', []))}]" if sp_info.get('frame_options') else ""
                st.markdown(f"**{'✅' if sp_deliv else '🔩'} {sp_info.get('equipment','')} ({sp_info.get('company','')})** │ 납기: {sp_dd} │ {sp_status_txt}")
                
                info_col1, info_col2 = st.columns(2)
                with info_col1:
                    st.write(f"**장비명:** {sp_info.get('equipment','')}")
                    st.write(f"**업체:** {sp_info.get('company','')}")
                    st.write(f"**납품예정:** {sp_dd}")
                    st.write(f"**상태:** {sp_status_txt}")
                with info_col2:
                    st.write(f"**진척률:** {sp_pct}%")
                    st.write(f"**검사점수:** {sp_score}/100")
                    st.write(f"**외관:** {sp_info.get('exterior_spec','')} │ **내부:** {sp_info.get('interior_spec','')}")
                    st.write(f"**파트:** {sp_info.get('frame_parts','')}덩어리{opts_str}")
                
                cat_progress = {}
                for item in CHECKLIST_ITEMS:
                    cat = item["category"]
                    if cat not in cat_progress: cat_progress[cat] = {"total": 0, "done": 0}
                    cat_progress[cat]["total"] += 1
                    key = str(item["no"])
                    if key in sp_checks and sp_checks[key].get("status") == "확인":
                        cat_progress[cat]["done"] += 1
                
                cat_cols = st.columns(len(cat_progress))
                for ci, (cat, data) in enumerate(cat_progress.items()):
                    with cat_cols[ci]:
                        cp = int(data["done"]/data["total"]*100) if data["total"]>0 else 0
                        st.metric(cat, f"{data['done']}/{data['total']}", f"{cp}%")
                
                if not sp_deliv and user["role"] == "admin":
                    if st.button("✅ 납품 완료 처리", key=f"cal_dv_dlg_{pid_item}", use_container_width=True):
                        st.session_state.projects[pid_item]["info"]["is_delivered"] = True
                        save_to_sheets(st.session_state.projects)
                        st.session_state.flash_msg = f"✅ [{sp_info.get('equipment')}] 납품 처리 완료!"
                        st.rerun()
                st.markdown("---")

        st.markdown(f"""
            <div style="font-size:34px; font-weight:800; color:#1e293b; margin-top:10px; margin-bottom:20px;">
                {get_logo_html('34px')} 프로젝트 납기 캘린더
            </div>
        """, unsafe_allow_html=True)
        st.markdown("달력에 표시된 **프로젝트 버튼을 클릭**하면 상세 현황 창이 나타납니다.")
        st.markdown("---")

        today = date.today()
        col_y, col_m, _ = st.columns([2, 2, 6])
        year = col_y.selectbox("년도 선택", range(today.year - 1, today.year + 3), index=1)
        month = col_m.selectbox("월 선택", range(1, 13), index=today.month - 1)

        month_str = f"{year}-{month:02d}"
        month_proj_map = {} 
        for pid, p in projects.items():
            dd = p.get('info', {}).get('delivery_date', '')
            if dd.startswith(month_str):
                if dd not in month_proj_map:
                    month_proj_map[dd] = []
                month_proj_map[dd].append((pid, p))

        cal = calendar.monthcalendar(year, month)
        
        # 요일 헤더
        day_names = ["월", "화", "수", "목", "금", "토", "일"]
        cols = st.columns(7)
        for i, day_name in enumerate(day_names):
            color = "#e74c3c" if day_name == "일" else "#3b82f6" if day_name == "토" else "#333"
            cols[i].markdown(f"<div style='text-align:center; font-weight:bold; color:{color}; padding:10px; background-color:#f8fafc; border:1px solid #e2e8f0; border-radius:5px; margin-bottom:5px;'>{day_name}</div>", unsafe_allow_html=True)
        
        # 🔥 리얼 달력 UI: 프로젝트 유무와 상관없이 고정된 높이(130px)로 렌더링 🔥
        for week in cal:
            cols = st.columns(7)
            for i, day in enumerate(week):
                with cols[i]:
                    with st.container(height=130, border=True):
                        if day == 0:
                            st.write("") # 빈칸 출력
                        else:
                            date_str = f"{year}-{month:02d}-{day:02d}"
                            is_today = (date_str == today.strftime("%Y-%m-%d"))
                            day_color = "#e74c3c" if i == 6 else "#3b82f6" if i == 5 else "#333"
                            weight = "900" if is_today else "bold"
                            
                            st.markdown(f'<div style="font-weight:{weight}; font-size:16px; margin-bottom:6px; color:{day_color};">{day}{" (오늘)" if is_today else ""}</div>', unsafe_allow_html=True)
                            
                            day_projs = month_proj_map.get(date_str, [])
                            for pid_item, p in day_projs:
                                info = p.get("info", {})
                                is_deliv = info.get("is_delivered", False)
                                
                                start_of_week = today - timedelta(days=today.weekday())
                                end_of_week = start_of_week + timedelta(days=6)
                                try: d_date = datetime.strptime(date_str, "%Y-%m-%d").date()
                                except: d_date = today
                                
                                # 색상 판별: 완료=회색, 지연/이번주내=빨간색, 일반=파란색
                                if is_deliv: 
                                    badge_color = "#95a5a6" 
                                elif d_date <= end_of_week: 
                                    badge_color = "#e74c3c" 
                                else: 
                                    badge_color = "#3b82f6" 
                                
                                # 마커 CSS 방식으로 캘린더 버튼만 완벽하게 색상 주입
                                marker = f"cal_mk_{pid_item}_{date_str.replace('-','')}"
                                st.markdown(f"<div class='{marker}' style='display:none;'></div>", unsafe_allow_html=True)
                                st.markdown(f"""
                                <style>
                                div.element-container:has(.{marker}) + div.element-container button,
                                div.stElementContainer:has(.{marker}) + div.stElementContainer button {{
                                    background-color: {badge_color} !important;
                                    border-color: {badge_color} !important;
                                    color: white !important;
                                    padding: 2px 5px !important;
                                    width: 100% !important;
                                    min-height: 0px !important;
                                    height: auto !important;
                                    border-radius: 4px !important;
                                    transition: filter 0.2s !important;
                                }}
                                div.element-container:has(.{marker}) + div.element-container button p {{
                                    color: white !important;
                                    font-size: 13px !important;
                                    font-weight: 700 !important;
                                    margin: 0 !important;
                                }}
                                div.element-container:has(.{marker}) + div.element-container button:hover,
                                div.stElementContainer:has(.{marker}) + div.stElementContainer button:hover {{
                                    filter: brightness(0.85) !important;
                                }}
                                </style>
                                """, unsafe_allow_html=True)
                                
                                # 클릭 가능한 프로젝트 버튼 (팝업 호출, 이모티콘 뺌)
                                if st.button(f"{info.get('equipment')}", key=f"cal_btn_{pid_item}_{date_str}", help=f"{info.get('company')} / 납기: {info.get('delivery_date')}", use_container_width=True):
                                    show_project_details_dialog(date_str, [(pid_item, p)])

        if user["role"] == "admin":
            st.markdown("<br><br>", unsafe_allow_html=True)
            st.subheader("납품 완료 간편 처리")
            st.markdown("달력에서 현황을 확인한 후 아래에서 납품 완료 상태로 즉시 변경할 수 있습니다.")
            
            pending_projs = {pid: p for pid, p in projects.items() if not p.get("info", {}).get("is_delivered", False)}
            if pending_projs:
                format_fn = lambda pid: f"{pending_projs[pid]['info']['equipment']} ({pending_projs[pid]['info']['company']} / 납기: {pending_projs[pid]['info']['delivery_date']})"
                selected_pid_to_complete = st.selectbox("완료 처리할 프로젝트를 선택하세요", list(pending_projs.keys()), format_func=format_fn)
                
                if st.button("선택한 프로젝트 납품 완료 처리", type="primary"):
                    st.session_state.projects[selected_pid_to_complete]["info"]["is_delivered"] = True
                    save_to_sheets(st.session_state.projects)
                    st.session_state.flash_msg = "✅ 납품 처리가 완료되었습니다."
                    st.rerun()
            else:
                st.info("현재 납품 대기 중인 프로젝트가 없습니다.")

    # ═══════════════════════════════════════════════════
    # ➕ 신규 등록
    # ═══════════════════════════════════════════════════
    elif menu == "➕ 신규 등록" and user["role"] == "admin":
        st.markdown(f"""
            <div style="font-size:34px; font-weight:800; color:#1e293b; margin-top:10px; margin-bottom:20px;">
                {get_logo_html('34px')} 신규 프로젝트 등록
            </div>
        """, unsafe_allow_html=True)
        st.markdown("---")

        with st.form("new_project_form", clear_on_submit=True):
            company = st.radio("업체명 *", ["한울산업", "정한테크"], horizontal=True)
            
            c1, c2 = st.columns(2)
            equipment = c1.text_input("장비명 (설비명) *", placeholder="장비명을 입력하세요")
            order_date = c2.date_input("발주일")
            
            c3, c4 = st.columns(2)
            frame_parts = c3.number_input("Frame Part 수 (덩어리) *", min_value=1, value=1, step=1)
            delivery_date = c4.date_input("납품 예정일자")

            st.markdown("<br><b>프레임 옵션</b>", unsafe_allow_html=True)
            opt_cols = st.columns(3)
            with opt_cols[0]: opt_clean = st.checkbox("클린부스")
            with opt_cols[1]: opt_table = st.checkbox("테이블")
            with opt_cols[2]: opt_jig = st.checkbox("전도방지지그")
            frame_options = [opt for opt, checked in zip(["클린부스", "테이블", "전도방지지그"], [opt_clean, opt_table, opt_jig]) if checked]

            st.markdown("<hr style='margin:15px 0;'>", unsafe_allow_html=True)
            
            spec_col1, spec_col2 = st.columns(2)
            with spec_col1: exterior_spec = st.radio("외관 사양", ["SUS", "도장"], horizontal=True)
            with spec_col2: interior_spec = st.radio("내부 사양", ["SUS", "도장"], horizontal=True)

            notes_top = st.text_area("특이사항", placeholder="추가 전달 사항을 입력하세요", height=80)
            
            st.markdown("<br>", unsafe_allow_html=True)
            submitted = st.form_submit_button("프로젝트 등록", use_container_width=True)

            if submitted:
                if not equipment: st.error("장비명은 필수 입력입니다.")
                else:
                    try:
                        pid = f"{company}_{equipment}_{datetime.now().strftime('%Y%m%d%H%M%S')}"
                        st.session_state.projects[pid] = repair_project({
                            "info": {
                                "company": company, "equipment": equipment, "order_date": str(order_date),
                                "delivery_date": str(delivery_date), "frame_parts": frame_parts,
                                "frame_options": frame_options, "exterior_spec": exterior_spec,
                                "interior_spec": interior_spec, "notes_top": notes_top,
                                "created_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
                                "delivery_delay_count": 0, "delay_total_biz_days": 0, "delay_request": {}, "is_delivered": False
                            },
                            "checks": {}, "special_notes": "", "history": []
                        })
                        save_to_sheets(st.session_state.projects)
                        st.session_state.flash_msg = "✅ 신규 프로젝트가 성공적으로 등록되었습니다!"
                        st.rerun()
                    except Exception as e:
                        st.error(f"등록 중 오류 발생: {e}")

    # ═══════════════════════════════════════════════════
    # 📋 점검
    # ═══════════════════════════════════════════════════
    elif menu == "📋 점검":
        if not projects: st.info("점검할 프로젝트가 없습니다.")
        else:
            active_projects = {pid: p for pid, p in projects.items() if not p.get("info", {}).get("is_delivered", False)}

            if st.session_state.inspection_project is None:
                st.markdown(f"""
                    <div style="font-size:34px; font-weight:800; color:#1e293b; margin-top:10px; margin-bottom:20px;">
                        {get_logo_html('34px')} 진행중인 프로젝트 점검
                    </div>
                """, unsafe_allow_html=True)
                
                if not active_projects:
                    st.success("모든 프로젝트가 납품 완료되었습니다!")
                
                for comp_name in ["한울산업", "정한테크"]:
                    comp_projs = [(k, v) for k, v in active_projects.items() if v.get("info", {}).get("company") == comp_name]
                    if comp_projs:
                        st.markdown(f'<div style="background:#f8f9fa; border-radius:12px; padding:16px 20px 6px 20px; margin-bottom:8px;"><div style="font-size:22px; font-weight:bold; color:#2c3e50; padding-bottom:8px; border-bottom:3px solid #4A90D9;">{comp_name}</div></div>', unsafe_allow_html=True)
                        for pid, proj in comp_projs:
                            pct = int(calc_progress(proj.get("checks", {})) * 100)
                            dd = proj.get("info", {}).get("delivery_date", "")
                            bar_c = ('#27ae60' if pct>=80 else '#4A90D9' if pct>=50 else '#f39c12' if pct>0 else '#e74c3c')
                            
                            col1, col2 = st.columns([8, 2])
                            with col1:
                                kanban_html = (
                                    f'<div style="background:white; border-radius:10px; padding:16px; margin-bottom:10px; border-left:5px solid #4A90D9; box-shadow:0 1px 6px rgba(0,0,0,0.07);">'
                                    f'<b style="font-size:18px;">{proj["info"]["equipment"]}</b> (납기: {dd})'
                                    f'<div style="background:#e8ecf1;border-radius:6px;height:10px;margin-top:8px;">'
                                    f'<div style="background:{bar_c};height:100%;width:{max(pct,2)}%;border-radius:6px;"></div>'
                                    f'</div></div>'
                                )
                                st.markdown(kanban_html, unsafe_allow_html=True)
                            with col2:
                                st.markdown("<br>", unsafe_allow_html=True)
                                if st.button("점검 시작", key=f"ins_{pid}", use_container_width=True):
                                    st.session_state.inspection_project = pid
                                    st.rerun()
            else:
                pid = st.session_state.inspection_project
                proj = projects.get(pid)
                if not proj:
                    st.session_state.inspection_project = None
                    st.rerun()

                info = proj.get("info", {})
                checks = proj.get("checks", {})
                
                if st.button("프로젝트 목록으로 돌아가기"):
                    st.session_state.inspection_project = None
                    st.rerun()

                delay_req = info.get("delay_request", {})
                st.markdown("---")
                st.markdown("### 일정 관리 및 조율")
                
                if user["role"] == "vendor":
                    if delay_req and delay_req.get("status") == "pending":
                        st.info(f"관리자에게 납기 변경을 요청했습니다. (희망일: {delay_req.get('requested_date')} / 승인 대기중)")
                    else:
                        with st.expander("관리자에게 납기 변경 요청하기"):
                            with st.form(f"delay_form_{pid}", clear_on_submit=True):
                                new_date = st.date_input("변경 희망 납품일")
                                reason = st.text_input("납기 변경(지연) 사유")
                                if st.form_submit_button("요청 전송", use_container_width=True):
                                    proj["info"]["delay_request"] = {"requested_date": str(new_date), "reason": reason, "status": "pending"}
                                    save_to_sheets(st.session_state.projects)
                                    st.session_state.flash_msg = "납기 변경 요청이 전송되었습니다."
                                    st.rerun()

                if user["role"] == "admin":
                    if delay_req and delay_req.get("status") == "pending":
                        st.warning(f"**[{info.get('company')}] 납기 변경 요청** : 기존 {info.get('delivery_date')} -> **희망 {delay_req.get('requested_date')}** (사유: {delay_req.get('reason')})")
                        col_a, col_b, col_c = st.columns([2, 2, 6])
                        if col_a.button("요청 승인 (납기 적용)", use_container_width=True):
                            old_dd = info.get("delivery_date", "")
                            new_dd = delay_req["requested_date"]
                            biz_days = count_business_days(old_dd, new_dd)
                            proj["info"]["delivery_date"] = new_dd
                            proj["info"]["delivery_delay_count"] = info.get("delivery_delay_count", 0) + 1
                            proj["info"]["delay_total_biz_days"] = info.get("delay_total_biz_days", 0) + biz_days
                            proj["info"]["delay_request"]["status"] = "approved"
                            save_to_sheets(st.session_state.projects)
                            st.session_state.flash_msg = "✅ 납기 변경 요청이 승인되었습니다."
                            st.rerun()
                        if col_b.button("요청 반려", use_container_width=True):
                            proj["info"]["delay_request"]["status"] = "rejected"
                            save_to_sheets(st.session_state.projects)
                            st.session_state.flash_msg = "납기 변경 요청이 반려되었습니다."
                            st.rerun()
                    
                    with st.expander("관리자 권한: 납기일 임의 변경 (지연 누적 적용)"):
                        with st.form(f"admin_delay_form_{pid}", clear_on_submit=True):
                            try: default_d = datetime.strptime(info.get("delivery_date", ""), "%Y-%m-%d").date()
                            except: default_d = date.today()
                            
                            new_date_admin = st.date_input("새로운 납기일로 변경", value=default_d)
                            st.warning("납기일을 변경하면 해당 업체의 '납기 지연 횟수'가 1회 증가합니다.")
                            confirm_change = st.checkbox("위 내용을 확인하였으며 납기일을 변경합니다.")
                            
                            if st.form_submit_button("납기일 변경 적용", use_container_width=True):
                                if str(new_date_admin) == info.get("delivery_date", ""):
                                    st.error("기존 납기일과 동일합니다. 다른 날짜를 선택하세요.")
                                elif not confirm_change:
                                    st.error("변경 사항을 적용하시려면 확인 체크박스를 선택해주세요.")
                                else:
                                    old_dd_str = info.get("delivery_date", "")
                                    biz_days = count_business_days(old_dd_str, str(new_date_admin))
                                    proj["info"]["delivery_date"] = str(new_date_admin)
                                    proj["info"]["delivery_delay_count"] = info.get("delivery_delay_count", 0) + 1
                                    proj["info"]["delay_total_biz_days"] = info.get("delay_total_biz_days", 0) + biz_days
                                    save_to_sheets(st.session_state.projects)
                                    st.session_state.flash_msg = "✅ 구글 시트에 납기일이 변경되었습니다!"
                                    st.rerun()

                st.markdown("---")
                opts_html = f'<span style="background:#eff6ff; color:#3b82f6; padding:4px 12px; border-radius:20px; font-size:14px; font-weight:700; border:1px solid #bfdbfe; margin-left:15px;">옵션: {", ".join(info.get("frame_options", []))}</span>' if info.get("frame_options") else ''

                st.markdown(f"""
                <div style="background:#eef2f7;padding:16px 22px;border-radius:10px;margin-bottom:24px;">
                    <div style="display:flex; align-items:center; margin-bottom:8px;">
                        <h3 style="margin:0;font-size:23px;">{info.get('equipment')}</h3>{opts_html}
                    </div>
                    <p style="margin:4px 0;font-size:17px;">
                        <b>업체:</b> {info.get('company')} │ <b>납품일:</b> <span style="color:#e74c3c;font-weight:bold;">{info.get('delivery_date')}</span> (납기변경 누적 {info.get('delivery_delay_count', 0)}회) │
                        <b>진척률:</b> {int(calc_progress(checks)*100)}%<br>
                        <b>외관:</b> {info.get('exterior_spec','')} │ <b>내부:</b> {info.get('interior_spec','')} │ <b>파트:</b> {info.get('frame_parts','')}덩어리
                    </p>
                </div>
                """, unsafe_allow_html=True)

                if user["role"] == "admin":
                    with st.expander("프로젝트 정보 수정"):
                        with st.form(f"edit_info_{pid}", clear_on_submit=False):
                            edit_c1, edit_c2 = st.columns(2)
                            with edit_c1:
                                edit_equipment = st.text_input("장비명", value=info.get("equipment", ""))
                                edit_company = st.radio("업체명", ["한울산업", "정한테크"], index=0 if info.get("company") == "한울산업" else 1, horizontal=True)
                                edit_parts = st.select_slider("Frame Part 수", options=list(range(1, 11)), value=int(info.get("frame_parts", 1)))
                            with edit_c2:
                                edit_ext = st.radio("외관 사양", ["SUS", "도장"], index=0 if info.get("exterior_spec") == "SUS" else 1, horizontal=True)
                                edit_int = st.radio("내부 사양", ["SUS", "도장"], index=0 if info.get("interior_spec") == "SUS" else 1, horizontal=True, key=f"edit_int_{pid}")
                            
                            st.markdown("**프레임 옵션**")
                            eo_cols = st.columns(3)
                            cur_opts = info.get("frame_options", [])
                            with eo_cols[0]: eo_clean = st.checkbox("클린부스", value="클린부스" in cur_opts, key=f"eo_c_{pid}")
                            with eo_cols[1]: eo_table = st.checkbox("테이블", value="테이블" in cur_opts, key=f"eo_t_{pid}")
                            with eo_cols[2]: eo_jig = st.checkbox("전도방지지그", value="전도방지지그" in cur_opts, key=f"eo_j_{pid}")
                            
                            edit_notes = st.text_area("특이사항", value=info.get("notes_top", ""), height=60)
                            
                            if st.form_submit_button("변경사항 저장", use_container_width=True):
                                proj["info"]["equipment"] = edit_equipment
                                proj["info"]["company"] = edit_company
                                proj["info"]["frame_parts"] = edit_parts
                                proj["info"]["exterior_spec"] = edit_ext
                                proj["info"]["interior_spec"] = edit_int
                                proj["info"]["frame_options"] = [o for o, c in zip(["클린부스", "테이블", "전도방지지그"], [eo_clean, eo_table, eo_jig]) if c]
                                proj["info"]["notes_top"] = edit_notes
                                save_to_sheets(st.session_state.projects)
                                st.session_state.flash_msg = "✅ 프로젝트 정보가 수정되었습니다!"
                                st.rerun()

                check_date = st.date_input("당일 점검 일자 지정", value=date.today(), key="cd")
                updated_checks = copy.deepcopy(checks)
                current_cat = ""

                for item in CHECKLIST_ITEMS:
                    key = str(item["no"])
                    if item["category"] != current_cat:
                        current_cat = item["category"]
                        st.markdown(f'<div style="background:#4A90D9; color:white; padding:10px 18px; border-radius:8px; margin:22px 0 10px 0; font-weight:bold; font-size:19px;">{current_cat}</div>', unsafe_allow_html=True)

                    existing = checks.get(key, {})
                    c1, c2, c3 = st.columns([5, 2, 3])
                    c1.markdown(f"**{item['no']}.** {item['item']}")
                    status = c2.radio(f"s{key}", ["미점검", "확인", "미비"], index=["미점검", "확인", "미비"].index(existing.get("status", "미점검")), horizontal=True, label_visibility="collapsed")
                    memo = c3.text_input(f"m{key}", value=existing.get("memo", ""), placeholder="비고 입력", label_visibility="collapsed")
                    updated_checks[key] = {"status": status, "date": str(check_date) if status != "미점검" else existing.get("date", ""), "memo": memo}

                st.markdown("---")
                special_notes = st.text_area("현장 특이사항 기록", value=proj.get("special_notes", ""), height=120)

                if st.button("점검 결과 저장", type="primary", use_container_width=True):
                    proj["checks"] = updated_checks
                    proj["special_notes"] = special_notes
                    if "history" not in proj: proj["history"] = []
                    proj["history"].append({"date": str(check_date), "progress": int(calc_progress(updated_checks) * 100), "score": calc_score(updated_checks), "saved_at": datetime.now().strftime("%Y-%m-%d %H:%M")})
                    save_to_sheets(st.session_state.projects)
                    st.success("✅ 구글 시트에 점검 결과가 저장되었습니다.")

                if proj.get("history"):
                    st.markdown("### 점검 히스토리")
                    st.dataframe(pd.DataFrame(proj["history"])[["date", "progress", "score"]].rename(columns={"date":"점검일", "progress":"진척률(%)", "score":"점수"}), use_container_width=True, hide_index=True)

    # ═══════════════════════════════════════════════════
    # 양식 추출
    # ═══════════════════════════════════════════════════
    elif menu == "📥 양식 추출":
        st.markdown(f"""
            <div style="font-size:34px; font-weight:800; color:#1e293b; margin-top:10px; margin-bottom:20px;">
                {get_logo_html('34px')} 엑셀 양식 일괄 추출
            </div>
        """, unsafe_allow_html=True)
        st.markdown("---")
        
        if not projects: st.info("추출할 프로젝트가 없습니다.")
        else:
            months = sorted(list(set(p.get('info', {}).get('delivery_date', '')[:7] for p in projects.values() if p.get('info', {}).get('delivery_date', ''))), reverse=True)
            col_filter, _ = st.columns([3, 7])
            selected_month = col_filter.selectbox("조회 기간 선택 (년/월)", ["전체"] + months)
            
            filtered_projs = {pid: p for pid, p in projects.items() if selected_month == "전체" or p.get('info', {}).get('delivery_date', '').startswith(selected_month)}
            
            if not filtered_projs:
                st.warning("해당 기간에 해당하는 프로젝트가 없습니다.")
            else:
                col_title, col_check = st.columns([6, 2])
                with col_title:
                    st.markdown("### 추출할 프로젝트를 선택하세요")
                with col_check:
                    select_all = st.checkbox("전체 선택", key="select_all_extract")
                
                df_data = []
                for pid, p in filtered_projs.items():
                    info = p.get('info', {})
                    df_data.append({
                        "선택": select_all,
                        "pid": pid,
                        "장비명": info.get('equipment', ''),
                        "업체": info.get('company', ''),
                        "납기일": info.get('delivery_date', ''),
                        "진척률": f"{int(calc_progress(p.get('checks', {}))*100)}%"
                    })
                df = pd.DataFrame(df_data)
                
                edited_df = st.data_editor(
                    df,
                    column_config={
                        "선택": st.column_config.CheckboxColumn("선택", default=False),
                        "pid": None
                    },
                    disabled=["장비명", "업체", "납기일", "진척률"],
                    hide_index=True,
                    use_container_width=True
                )
                
                selected_pids = edited_df[edited_df["선택"] == True]["pid"].tolist()
                
                if selected_pids:
                    today_str = date.today().strftime('%Y%m%d')
                    
                    st.success(f"총 {len(selected_pids)}개 선택됨")
                    st.markdown("")
                    
                    if len(selected_pids) == 1:
                        proj = projects.get(selected_pids[0])
                        if proj:
                            excel_data = generate_checklist_excel(proj)
                            filename = f"Frame_점검표_{proj.get('info', {}).get('company', '')}_{proj.get('info', {}).get('equipment', '')}_{today_str}.xlsx"
                            st.download_button(
                                label=f"📥 {proj.get('info', {}).get('equipment', '')} 점검표 다운로드",
                                data=excel_data.getvalue(),
                                file_name=filename,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml",
                                key="dl_single",
                                use_container_width=True
                            )
                    else:
                        zip_buffer = io.BytesIO()
                        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
                            for pid in selected_pids:
                                proj = projects.get(pid)
                                if not proj: continue
                                excel_data = generate_checklist_excel(proj)
                                filename = f"Frame_점검표_{proj.get('info', {}).get('company', '')}_{proj.get('info', {}).get('equipment', '')}_{today_str}.xlsx"
                                zf.writestr(filename, excel_data.getvalue())
                        zip_buffer.seek(0)
                        
                        st.download_button(
                            label=f"📥 {len(selected_pids)}개 점검표 일괄 다운로드 (ZIP)",
                            data=zip_buffer.getvalue(),
                            file_name=f"Frame_점검표_일괄_{today_str}.zip",
                            mime="application/zip",
                            key="dl_zip_all",
                            use_container_width=True
                        )
                else:
                    st.info("위 표에서 다운로드할 프로젝트의 '선택' 칸을 체크해주세요.")
